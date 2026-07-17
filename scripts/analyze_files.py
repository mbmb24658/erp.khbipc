#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Comprehensive analyzer for the user's uploaded project files.

Handles 5 files in ./upload/:
    1. back.json    - Django backend specifications (template-style blocks)
    2. db.json      - Django fixture / sample data
    3. front.json   - Frontend (page/component) specifications
    4. MASTER_R05.xlsx - Excel workbook (to be imported into the DB)
    5. Book1.xlsx   - Excel workbook (auxiliary)

The JSON files appear to use a `key : {{ ... }}` template-block syntax
rather than a single JSON document, so we parse each block separately.
Pure-JSON files (db.json) are parsed with the standard json module first
and we fall back to block parsing if that fails.

Output is printed to stdout in clearly delimited sections.
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


UPLOAD_DIR = Path("./upload")
BACK_JSON = UPLOAD_DIR / "back.json"
DB_JSON = UPLOAD_DIR / "db.json"
FRONT_JSON = UPLOAD_DIR / "front.json"
MASTER_XLSX = UPLOAD_DIR / "MASTER_R05.xlsx"
BOOK1_XLSX = UPLOAD_DIR / "Book1.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def print_section(title: str, char: str = "=") -> None:
    line = char * 78
    print(f"\n{line}\n{title}\n{line}")


def print_subsection(title: str, char: str = "-") -> None:
    line = char * 78
    print(f"\n{line}\n{title}\n{line}")


def parse_blocks(text: str) -> list[tuple[str, dict]]:
    """
    Parse a file containing one or more blocks of the form:

        block_name : {{
            ... json object content (without the outer { }) ...
        }}

    or (with spaces, mixed style):

        block name with spaces (XX) : { {
            ...
        } }

    The double braces are a templating quirk; the actual JSON object body
    lives between them and we wrap it in `{ ... }` to recover a single
    JSON object per block.

    Block names may contain Persian/Arabic characters, spaces, parentheses
    and digits. Opening line ends with `{{` or `{ {` (with optional inner
    whitespace). Closing line is `}}` or `} }` (with optional inner
    whitespace).

    Returns a list of (block_name, parsed_object) tuples.
    """
    blocks: list[tuple[str, dict]] = []
    # Opening line: anything, then " : " then either "{{" or "{ {" at end.
    # Closing line: starts with "}}" or "} }" (possibly trailing whitespace).
    open_pat = re.compile(r"^(?P<name>.+?)\s*:\s*\{\s*\{\s*$")
    close_pat = re.compile(r"^\}\s*\}\s*$")

    lines = text.splitlines()
    i = 0
    n = len(lines)
    while i < n:
        m = open_pat.match(lines[i])
        if not m:
            i += 1
            continue
        name = m.group("name").strip()
        # Collect body until we hit the closing line
        body_lines = []
        j = i + 1
        found_close = False
        while j < n:
            if close_pat.match(lines[j]):
                found_close = True
                break
            body_lines.append(lines[j])
            j += 1
        if not found_close:
            print(f"  [WARN] Block '{name}' has no closing '}}' - skipped")
            i += 1
            continue
        body = "\n".join(body_lines).strip()
        try:
            obj = json.loads("{" + body + "}")
        except json.JSONDecodeError as e:
            print(f"  [WARN] Could not parse block '{name}': {e}")
            i = j + 1
            continue
        blocks.append((name, obj))
        i = j + 1
    return blocks


def parse_json_file(path: Path) -> tuple[dict | list | None, list[tuple[str, dict]]]:
    """
    Try to parse `path` as pure JSON first; if that fails, parse as blocks.
    Returns (pure_json_result_or_None, list_of_blocks).
    """
    text = path.read_text(encoding="utf-8")
    # Attempt 1: pure JSON
    try:
        return json.loads(text), []
    except json.JSONDecodeError:
        pass
    # Attempt 2: block-style
    blocks = parse_blocks(text)
    return None, blocks


# ---------------------------------------------------------------------------
# 1. back.json
# ---------------------------------------------------------------------------
def analyze_back_json() -> None:
    print_section("1. back.json - Django Backend Specifications")
    _, blocks = parse_json_file(BACK_JSON)
    print(f"Total blocks found: {len(blocks)}\n")

    all_models: list[dict] = []
    total_relationships: list[dict] = []

    for block_name, obj in blocks:
        module = obj.get("module", {})
        models = obj.get("models", [])
        apis = obj.get("api_endpoints", []) or obj.get("apis", [])
        print_subsection(
            f"Block '{block_name}' -> Module '{module.get('app_name')}'"
        )
        print(f"  module.id          : {module.get('id')}")
        print(f"  module.name (FA)   : {module.get('name')}")
        print(f"  module.description : {module.get('description')}")
        print(f"  module.version     : {module.get('version')}")
        print(f"  module.app_name    : {module.get('app_name')}")
        print(f"  module.dependencies: {module.get('dependencies')}")
        print(f"  #models            : {len(models)}")
        print(f"  #api_endpoints     : {len(apis)}")

        for m in models:
            name = m.get("name")
            app = m.get("app_label")
            fields = m.get("fields", [])
            print(f"\n    Model: {name}  (app_label={app}, #fields={len(fields)})")
            key_fields = []
            relationships = []
            for f in fields:
                fname = f.get("name")
                ftype = f.get("type")
                if ftype in ("ForeignKey", "OneToOneField", "ManyToManyField"):
                    relationships.append(
                        {
                            "field": fname,
                            "type": ftype,
                            "to": f.get("to"),
                            "related_name": f.get("related_name"),
                        }
                    )
                # Capture key scalar fields for the summary
                if fname in (
                    "id",
                    "wbs_code",
                    "title",
                    "name",
                    "code",
                    "level",
                    "parent",
                    "start_date",
                    "finish_date",
                    "progress_plan",
                    "progress_actual",
                    "status",
                    "amount",
                    "cost",
                    "weight",
                    "unit",
                    "category",
                    "description",
                    "date",
                    "owner",
                    "user",
                    "project",
                    "activity",
                    "resource",
                ) or f.get("primary_key"):
                    key_fields.append(f"{fname}:{ftype}")
            print(f"      key fields    : {', '.join(key_fields)}")
            if relationships:
                print(f"      relationships :")
                for r in relationships:
                    print(
                        f"        - {r['field']} ({r['type']}) -> {r['to']}"
                        + (f"  [related_name={r['related_name']}]" if r.get('related_name') else "")
                    )
                total_relationships.append(
                    {"model": name, "app": app, "rels": relationships}
                )
            all_models.append({"model": name, "app": app, "fields": fields})

    print_subsection("Summary: All modules")
    print(f"  Total modules: {len(blocks)}")
    print(f"  Total models : {len(all_models)}")
    app_counts = Counter(m["app"] for m in all_models)
    for app, cnt in app_counts.items():
        print(f"    - app '{app}': {cnt} model(s)")
    print(f"\n  Total cross/inner model relationships: {sum(len(r['rels']) for r in total_relationships)}")


# ---------------------------------------------------------------------------
# 2. db.json
# ---------------------------------------------------------------------------
def analyze_db_json() -> None:
    print_section("2. db.json - Django Fixture / Sample Data")
    pure, blocks = parse_json_file(DB_JSON)
    if pure is not None:
        print(f"Parsed as pure JSON. Top-level type: {type(pure).__name__}")
        if isinstance(pure, list):
            print(f"Total records: {len(pure)}")
            counts = Counter(rec.get("model", "<unknown>") for rec in pure)
            print_subsection("Record counts per model")
            for model, cnt in counts.most_common():
                print(f"  {model:50s} : {cnt}")
            # Show one sample record per model
            print_subsection("One sample record per model (truncated)")
            seen = set()
            for rec in pure:
                model = rec.get("model")
                if model in seen:
                    continue
                seen.add(model)
                fields = rec.get("fields", {})
                preview = json.dumps(fields, ensure_ascii=False)
                if len(preview) > 400:
                    preview = preview[:400] + " ... (truncated)"
                print(f"\n  model={model}  pk={rec.get('pk')}")
                print(f"    fields: {preview}")
        elif isinstance(pure, dict):
            print(f"Top-level dict keys: {list(pure.keys())[:20]}")
            # Possibly a dict of model -> list
            print_subsection("Keys and counts")
            for k, v in pure.items():
                cnt = len(v) if isinstance(v, (list, dict)) else 1
                print(f"  {k:40s} : {cnt} item(s)")
    else:
        print(f"Parsed as blocks. Total blocks: {len(blocks)}")
        for name, obj in blocks:
            print_subsection(f"Block '{name}'")
            if isinstance(obj, list):
                print(f"  list of {len(obj)} items")
                if obj:
                    print(f"  first item keys: {list(obj[0].keys())[:10]}")
            elif isinstance(obj, dict):
                print(f"  dict keys: {list(obj.keys())[:20]}")
                if "models" in obj:
                    for m in obj["models"]:
                        cnt = len(m.get("records", []))
                        print(f"    - {m.get('name', m.get('model'))}: {cnt} records")


# ---------------------------------------------------------------------------
# 3. front.json
# ---------------------------------------------------------------------------
def analyze_front_json() -> None:
    print_section("3. front.json - Frontend Specifications")
    _, blocks = parse_json_file(FRONT_JSON)
    print(f"Total blocks found: {len(blocks)}\n")

    total_pages = 0
    total_components_used = Counter()

    for block_name, obj in blocks:
        module = obj.get("module", {})
        routes = obj.get("routes", [])
        pages = obj.get("pages", [])
        components = obj.get("components", []) or obj.get("shared_components", [])
        admin = obj.get("admin", obj.get("admin_interface"))

        print_subsection(
            f"Block '{block_name}' -> Module '{module.get('name')}'"
        )
        print(f"  module.id          : {module.get('id')}")
        print(f"  module.name (FA)   : {module.get('name')}")
        print(f"  module.description : {module.get('description')}")
        print(f"  module.base_path   : {module.get('base_path')}")
        print(f"  module.dependencies: {module.get('dependencies')}")
        print(f"  #routes            : {len(routes)}")
        print(f"  #pages             : {len(pages)}")
        print(f"  #components        : {len(components)}")

        # Routes
        if routes:
            print(f"\n  Routes:")
            for r in routes:
                print(
                    f"    {r.get('path',''):20s} -> {r.get('component','')}"
                    + (" (exact)" if r.get("exact") else "")
                )

        # Pages
        for p in pages:
            total_pages += 1
            print(f"\n    Page: {p.get('name')}  (path={p.get('path')})")
            print(f"      description : {p.get('description')}")
            print(f"      components  : {p.get('components')}")
            for c in p.get("components", []) or []:
                total_components_used[c] += 1
            state = p.get("state")
            if state:
                print(f"      state keys  : {list(state.keys()) if isinstance(state, dict) else state}")
            actions = p.get("actions")
            if actions:
                if isinstance(actions, dict):
                    print(f"      actions     : {list(actions.keys())}")
            apis = p.get("api_calls")
            if apis:
                print(f"      api_calls   : {[a.get('name') for a in apis]}")

        # Components defined inline
        for c in components:
            print(f"\n    Component: {c.get('name')}")
            print(f"      description : {c.get('description')}")
            props = c.get("props", [])
            if props:
                print(f"      props       : {[p.get('name') for p in props]}")

        # Admin interface hints
        if admin:
            print(f"\n    Admin interface: {json.dumps(admin, ensure_ascii=False)}")

    print_subsection("Summary: Frontend")
    print(f"  Total modules : {len(blocks)}")
    print(f"  Total pages   : {total_pages}")
    print(f"  Most-used components:")
    for comp, cnt in total_components_used.most_common(15):
        print(f"    {comp:30s} : {cnt}")


# ---------------------------------------------------------------------------
# 4 & 5. Excel inspection
# ---------------------------------------------------------------------------
def cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def analyze_workbook(path: Path, label: str) -> None:
    print_section(f"{label}: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print_subsection(f"Sheet: '{sheet_name}'")
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        print(f"  dimensions : {ws.dimensions}")
        print(f"  max_row    : {max_row}")
        print(f"  max_column : {max_col}")

        if max_row == 0 or max_col == 0:
            print("  (empty sheet)")
            continue

        # Read all rows into a list (cap for safety)
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_row:
                break
            rows.append(row)

        # Detect header row - assume row 0 is header, but also check row 1
        # in case there's a title row.
        header_row_index = 0
        for i in range(min(3, len(rows))):
            non_empty = sum(1 for v in rows[i] if v not in (None, ""))
            if non_empty >= max_col * 0.6:
                header_row_index = i
                break

        headers = [cell_to_str(v) for v in rows[header_row_index]]
        print(f"\n  Header row (row {header_row_index + 1}):")
        for idx, h in enumerate(headers, 1):
            print(f"    Col {idx:>2} ({get_column_letter(idx)}): {h!r}")

        # First 3 data rows after header
        data_start = header_row_index + 1
        print(f"\n  First 3 data rows (starting at row {data_start + 1}):")
        for r_offset in range(3):
            r_idx = data_start + r_offset
            if r_idx >= len(rows):
                break
            row = rows[r_idx]
            print(f"\n    Row {r_idx + 1}:")
            for col_idx, h in enumerate(headers):
                val = row[col_idx] if col_idx < len(row) else None
                val_str = cell_to_str(val)
                if len(val_str) > 80:
                    val_str = val_str[:80] + " ...(truncated)"
                print(f"      {h!r:40s} = {val_str!r}")

        # Data row count (excluding header)
        data_rows = len(rows) - data_start
        # Trim trailing fully-empty rows
        trailing_empty = 0
        for r in reversed(rows[data_start:]):
            if all(v in (None, "") for v in r):
                trailing_empty += 1
            else:
                break
        effective_data_rows = data_rows - trailing_empty
        print(f"\n  Effective data rows (excl. trailing empties): {effective_data_rows}")

        # Column-wise non-empty count (helps find sparse columns)
        print(f"\n  Column fill statistics (non-empty count / {effective_data_rows}):")
        for col_idx, h in enumerate(headers):
            if not h:
                continue
            cnt = 0
            for r in rows[data_start : data_start + effective_data_rows]:
                if col_idx < len(r) and r[col_idx] not in (None, ""):
                    cnt += 1
            print(f"    {h!r:40s} : {cnt}")

    wb.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print_section("FILE ANALYSIS REPORT", "#")
    print(f"Upload dir: {UPLOAD_DIR}")
    for f in [BACK_JSON, DB_JSON, FRONT_JSON, MASTER_XLSX, BOOK1_XLSX]:
        size = f.stat().st_size
        print(f"  - {f.name:25s} {size:>10,} bytes  exists={f.exists()}")

    analyze_back_json()
    analyze_db_json()
    analyze_front_json()
    analyze_workbook(MASTER_XLSX, "4. MASTER_R05.xlsx")
    analyze_workbook(BOOK1_XLSX, "5. Book1.xlsx")

    print_section("END OF REPORT", "#")


if __name__ == "__main__":
    main()
